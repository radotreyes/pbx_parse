import settings, savedata, main
from savedata import Presets
from relay import *
from main import *

import re, os, json, sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, PatternFill

class RawFile():
    def __init__( self, path, book, request = None ):
        self.book = book
        self.name = os.path.split( path )[1].split( '.' )[0]
        self.request = Presets.request = request
        update( self.request, 'Opening file {}...'.format( path ) )

        with open( path ) as file:
            self.content = file.readlines()

        self.records = self.get_records()
        if self.records:
            for key, value in self.records.items():
                self.process_record( key, value )

    def get_records( self ):
        '''
        Scans raw file content and separates data records by looking for the string "COMMAND: LI". Data records are saved to _self.records_.
        '''
        records = {}; cache = []
        found = False; num = 0
        for i, line in enumerate( self.content ):
            # scan for EOF
            try:
                nextline = self.content[i+1]
            except IndexError:
                nextline = None

            # if we find a PBX command
            if re.search( r'^COMMAND: LI \w+', line ):
                # indicate that we have found a record
                found = True

                # if any data is currently cached
                if cache:
                    # save it to record list
                    key = self.name + '({})'.format( num ) if num else self.name
                    records[key] = cache
                    num += 1 # increment record number
                    cache = [] # clear cache

            # while in a record, save lines to cache
            if found:
                cache.append( line )

            # we've reached EOF, save cache to record list
            if not nextline:
                key = self.name + '({})'.format( num ) if num else self.name
                records[key] = cache

        return records

    def process_record( self, name, rawdata ):
        # parses the command associated with this data record
        cmd = rawdata[0].split( 'COMMAND:' )[1].split( '\n' )[0]

        # creates a new Record object from this data record
        test_record = Record( name, rawdata, cmd, self.book, self.request )


class Record():
    '''
    Data contained in a RawFile, as separated by a "COMMAND: LI". All relevant parsing guidelines, including structure, etc.
    are set on initialization.
    '''

    def __init__( self, name, rawdata, command, book, request ):
        # Open file
        run = True
        self.raw = rawdata
        self.request = request

        ''' LOADING PRESETS '''
        if Presets.pdata:
            while True:
                if self.request:
                    use = choose( 'You have saved presets.\nWould you like to load a preset for this file? (Y/N)' )
                else:
                    choose( 'You have saved presets.\nWould you like to load a preset for this file? (Y/N)' )
                    use = input( '>> ' ).lower()
                if true( use ):
                    pdata = Presets.get_pdata()
                    if pdata:
                        self.meta = pdata['meta']
                        self.structure = pdata['structure']
                        self.presets = True
                    else:
                        self.presets = True
                        run = False
                    break
                elif false( use ):
                    self.presets = False
                    break
                else:
                    confirm( 'Please enter a valid input.' )
        else:
            self.presets = False

        if not self.presets:
            # Metadata
            self.meta = {
                'name': name,
                'pk': None,
                'pk_inline': False, # if the PK contains an inline data point
                'command': command
            }
        if run:
            run = self.set_meta()

        if run:
            if not self.presets:
                self.structure = {
                    'keys': [], # contains the line of data above each horizontal rule
                                # used to verify uniqueness of each hrule pattern
                    'cells': [], # unique horizontal rule patterns in the file
                    'fix': [], # specifies if fields need to be align-fixed
                    'field_names': [], # user-defined field names
                    'set_ids': [], # integers corresponding 1:1 to a data pattern
                    'write_columns': [], # starting indices of grouped field names, 0ind
                }
            run = self.set_structure()

            if run:
                # Data object
                self.data = {} # actually contains the data

                # Output worksheet
                self.wb = { 'book': load_workbook( book ), 'path': book }
                self.sheet = self.wb['book'].create_sheet( self.meta['name'] )
                self.scrape()
                self.write()
            else:
                update( self.request, 'Ready' )
        else:
            update( self.request, 'Ready' )

    def __str__( self ):
        return 'Data file with name: ' + self.name

    def set_meta( self ):
        os.system( 'cls' if os.name == 'nt' else 'clear' )
        confirm( 'This program requires you to enter a "Primary Key" for the data set you wish to parse. A Primary Key is a unique data field that is present in every entry in a given data set.\n\nPlease note that Siemens PBX files may have multiple data sets in one file; you may be required to enter multiple Primary Keys for a given file.')
        while True:
            pk = prompt( 'Please enter the Primary Key for this data set.\n\nNOTE: Make sure that the Primary Key you are entering is unique and present\nin each entry in this data set. If incorrect, this program will not work correctly.' )
            if not pk or not re.search( r'\S', pk ):
                if pk == None:
                    return False
                else:
                    confirm( '\nPlease enter a non-blank name.\n' )
            else:
                break
        self.meta['pk'] = r'' + re.escape( pk )
        return True

        os.system( 'cls' if os.name == 'nt' else 'clear' )

    def set_structure( self ):
        '''
        - Assign each data field plus group names to a dict "structure".
        - Use the pattern of the horizontal rules
            PLUS the line immediately above each line of horizontal rules
            to uniquely identify each row of data.
        '''

        sample_data = {
            'locs': [],
            'hrules': [],
            'entries': []
        }

        point = False; f_id = 0
        for i, line in enumerate( self.raw ):
            try:
                nextline = self.raw[i+1]
            except IndexError:
                nextline = None

            if Scanner.is_pk( self.meta['pk'], line ): # upon finding the PK...
                point = not point # we've entered or exited a data point

                # If the PK contains a colon, then include the text after that colon as a data point
                if re.search( r':', line ) and not self.meta['pk_inline']:
                    self.structure['keys'].append( 'PRIMARY KEY' )
                    self.structure['cells'].append( [ (
                        len( line.split( ':' )[0] ) + 1,
                        len( line ) - 1
                    ), ] )
                    self.structure['set_ids'].append( f_id )
                    self.structure['fix'].append( False )
                    self.meta['pk_inline'] = True
                    f_id += 1

            if point and Scanner.is_hrule( line, nextline ): # if we find a horizontal rule
                ( k, c ) = Scanner.get_pattern( self.raw, i )

                # store the horizontal rule pattern
                if k not in self.structure['keys']:
                    self.structure['keys'].append( k )
                    self.structure['cells'].append( c )
                    self.structure['set_ids'].append( f_id )

                    # does the data need to be align-fixed?
                    if Scanner.is_data( self.raw[i+1] ):
                        '''
                        Data can sometimes be shifted one line down.
                        If this happens, an align-fix is needed.
                        If an align-fix is needed, data will be parsed from
                        the next line down, rather than the line that the data
                        is normally on.
                        '''
                        if re.search( self.raw[i+1][2:].strip(), self.raw[i+2] ):
                            self.structure['fix'].append( True )
                        else:
                            self.structure['fix'].append( False )

                    f_id += 1

                    ''' DISPLAYED IN UI '''
                    sample_data['locs'].append( str( i ) )
                    sample_data['hrules'].append( self.raw[i] )
                    sample_data['entries'].append( self.raw[i+1] )
        if not self.structure['cells'] or len( self.structure['keys'] ) != len( self.structure['cells'] ):
            confirm( 'Couldn\'t process this file. Did you enter a valid Primary Key?' )
            return False

        if not self.presets:
            ''' UI '''
            os.system( 'cls' if os.name == 'nt' else 'clear' )

            x = offset = 0 # offset index, to be used if pk_inline is True
            for i, key in enumerate( self.structure['keys'] ):
                x = i + offset

                self.structure['field_names'].append( [] )

                for j, cell in enumerate( self.structure['cells'][i] ):
                    if key != 'PRIMARY KEY':
                        ( f, l ) = cell

                        message = '> ' + key + '\n'
                        message += '> ' + sample_data['hrules'][x] + '\n'
                        message += '> ' + sample_data['entries'][x] + '\n'
                        message += '> ' + ' ' * int( f ) + '^' + '\n'
                        message += '> ' + ' ' * int( f ) + '^' + '\n'

                        confirm( 'Please provide names for each field in the file "{}".\n\nAll fields must be properly named to store data from this raw file. If necessary, please browse through the raw file to determine appropriate names.'.format( self.meta['name'] ) )

                        while True:
                            lp = LongPrompt(
                                'Name data files',
                                'Please provide a name for the data field shown below.\nNames must consist of non-blank characters.\nData fields which are on the same line must have unique names.\n\nDisplaying data from LINE ' + sample_data['locs'][x] + ':',
                                message
                             )
                            self.request.wait_window( lp.top )
                            name = lp.response.get()

                            if not name or not re.search( r'\S', name ):
                                confirm( '\nPlease enter a non-blank name.\n' )
                            elif name in self.structure['field_names'][i]:
                                confirm( '\nThat name is already in use for this line. Please enter a unique name.\n' )
                            elif name == 'ABORT':
                                return False
                            else:
                                break
                    else:
                        name = self.meta['pk']
                        offset -= 1

                    self.structure['field_names'][i].append( name )
            ''' /UI '''
        update(
            self.request, 'Scanning file {}...'
                .format( self.meta['name']
            )
        )

        ''' SAVING PRESETS '''
        while True:
            if self.request:
                save = choose( 'Save your entries as a new preset? (Y/N)' )
            else:
                choose( 'Save your entries as a new preset? (Y/N)' )
                save = input( '>> ' ).lower()
            if true( save ):
                Presets.append_pdata( self.meta, self.structure )
                Presets.save_pdata()
                break
            elif false( save ):
                break
            else:
                confirm( 'Please enter a valid input.' )

        os.system( 'cls' if os.name == 'nt' else 'clear' )
        return True


    def scrape( self ):
        '''
        Look for data according to self.structure and store it in _self.data_.
        '''
        n = -1 # member index
        for i, line in enumerate( self.raw ):
            try:
                nextline = self.raw[i+1]
            except IndexError:
                nextline = None

                # push the last data_set to member
                try:
                    member['**SET_ID: ' + set_key] = data_set
                except UnboundLocalError: pass

                # push the last member to self.data
                self.data[n] = member

            if Scanner.is_pk( self.meta['pk'], line ):
                # create a new data member
                if n != -1:
                    # push previous data set into current member
                    try:
                        member['**SET_ID: ' + set_key] = data_set
                    except UnboundLocalError: pass
                    # carry on in case we haven't found a data_set yet

                    ''' zip the previously parsed data point'''
                    self.data[n] = member

                member = {}
                try: data_set = {}
                except UnboundLocalError: pass

                n += 1

                if self.meta['pk_inline']:
                    i = self.structure['cells'][0][0][0]
                    j = len( line ) - 1
                    member['**SET_ID: 0'] = {
                        '**ENTRY#: 0': {
                            self.meta['pk']: line[i:j].strip()
                        }
                    }

            if Scanner.is_hrule( line, nextline ):
                # push previous data set into current member
                try:
                    member['**SET_ID: ' + set_key] = data_set
                except UnboundLocalError:
                    pass # carry on in case we haven't found a data_set yet

                # retrieve the structure index determined by the closest key
                for j, key in enumerate( self.structure['keys'] ):
                    if self.raw[i-1] == key:
                        index = j

                # grab field data from self.structure
                names_to_map = self.structure['field_names'][index]
                cells_to_map = self.structure['cells'][index]
                id_to_map = self.structure['set_ids'][index]
                fix_to_map = self.structure['fix'][index]

                # sort each set into its own dict
                set_key = str( id_to_map )

                # if there is already an entry in this set under the same hrule pattern:
                try:
                    if member['**SET_ID: ' + set_key]:
                        data_set = member['**SET_ID: ' + set_key]
                        c = len( data_set )
                    else:
                        c = 0
                        data_set = {}
                except KeyError:
                    c = 0
                    data_set = {}

            if Scanner.is_data( line ):
                this_names = names_to_map

                if fix_to_map:
                    '''
                    This path applies an ALIGN-FIX.
                        - Instead of parsing the line the data is on,
                            parse the next line shifted by 2.
                        - Sometimes necessary in faulty Siemens PBX
                            output files.
                    '''
                    this_data = [
                        Scanner.parse( 'DS' + self.raw[i+1], cell )
                        for cell in cells_to_map
                    ]
                else:
                    this_data = [
                        Scanner.parse( line, cell )
                        for cell in cells_to_map
                    ]

                # push data into appropriate set
                data_set['**ENTRY#: {}'.format( c )] = dict( zip( this_names, this_data ) )

                c += 1 # number of times data has been pushed under one hrule

    def write( self ):
        '''
        Write the file to Excel.
        '''
        update( self.request, 'Writing to workbook...' )
        flat_fields = [] # flattened list of field names for column assignment
        for group in self.structure['field_names']:
            for i, field in enumerate( group ):
                flat_fields.append( field )
                if i == 0:
                    self.structure['write_columns'].append( len( flat_fields ) - 1 )

        for i, field in enumerate( flat_fields ):
            self.sheet.cell( row = 1, column = i+1 ).value = field
            self.sheet.cell( row = 1, column = i+1 ).fill =     PatternFill( 'solid', fgColor = "D0D0D0" )

        current_row = 2
        for i, m in enumerate( self.data ):
            current_row += Scanner.transcribe( self.data[m], self.sheet, current_row, self.structure['write_columns'] )

        self.wb['book'].save( self.wb['path'] )

        ''' DEBUG'''
        confirm( 'Parsing finished.\nFile was saved to workbook {}.'.format( self.wb['path'] ) )
        update( self.request, 'Ready' )

class Scanner():
    @staticmethod
    def is_hrule( line, nextline ):
        ''' Number of spaced horizontal rules defines the number of fields for a given member '''
        if nextline:
            return re.search( r'-+', line ) and not re.match( r'^DS', line ) and re.match( r'^DS', nextline )

        return False

    @staticmethod
    def is_pk( pk, line ):
        '''
        Searches for a primary key in the given line.
        IN: _pk_, primary key to look for
            _line_, line to search in
        '''
        return re.search( pk, line )

    @staticmethod
    def get_pattern( hrule_line, index ):
        '''
        Grabs the horizontal rule pattern from the appropriate line.
        IN: _hrule_line_, a line a hrule pattern
            _index_, the index of that line
        '''
        p = re.finditer( r'-+', hrule_line[index] )
        this_key = hrule_line[index - 1]
        this_pattern = []

        while True:
            try:
                this_pattern.append( next( p ).span() )
            except StopIteration: break

        return ( this_key, this_pattern )

    @staticmethod
    def is_data( line ):
        '''
        If "DS" is present at the start of a line, then that line contains data
        IN: _line_, the line to scan
        '''
        return re.match( r'^DS', line )

    @staticmethod
    def parse( line, cell ):
        '''
        Grabs data from a line according to its data structure,
        which is defined in _FileObject.structure_
        IN: _line_, the line to scan
            _cell_, the coordinate span of _line_ which contains data
        '''
        ( i, j ) = cell
        return line[i:j].strip() if re.match( r'\S', line[i:j] ) else ' '

    @staticmethod
    def transcribe( member, ss, cur, wc ):
        '''
        Writes data to Excel.
        IN: _member_, the data member which is being transcribed
            _ss_, the spreadsheet to write to
            _cur_, the current row number which should be written to
            _wc_, write column; the column number which should be written to
        '''
        set_len = 1 # rows to be reserved for this particular data set
        for s in member: # find the number of rows required
            set_id = int( re.search( r'\d+', s ).group() )
            set_len = len( member[s] ) if len( member[s] ) > set_len else set_len

            for e in member[s]:
                entry_num = int( re.search( r'\d+', e ).group() )
                for i, f in enumerate( member[s][e] ):
                    r = cur + entry_num
                    c = wc[set_id] + i + 1

                    cell = ss.cell( row = r, column = c )
                    cell.value = member[s][e][f]
                    if c == 1:
                        cell.fill = PatternFill( 'solid', fgColor = "FFFF00" )

        return set_len

if __name__ == '__main__':
    ''' testing presets '''
    Presets.load_pfile()

    ''' for use with test file '''
    book = Workbook()
    file_rp_all = RawFile( 'EXTEN.txt', 'test.xlsx' )
